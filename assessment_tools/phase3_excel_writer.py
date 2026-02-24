# phase3_excel_writer.py

import shutil
from datetime import datetime
from pathlib import Path
import openpyxl

TEMPLATE_FILE  = "地域防災計画確認票.xlsm"
SHEET_NAME     = "Sheet1"
HEADER_ROW     = 2   # ヘッダー行（質問No / 確認事項 / 回答）
DATA_START_ROW = 3   # 質問データ開始行
COL_QID        = 2   # B列: 質問No
COL_ANSWER     = 4   # D列: 回答
COL_EVIDENCE   = 5   # E列: 根拠ページ（追記）


def write_answers_to_excel(answers: dict, template_path: str = TEMPLATE_FILE) -> str:
    """
    回答辞書をExcelテンプレートに書き込み、出力ファイルパスを返す。

    Args:
        answers: { qid(int): {"answer": str, "evidence_pages": list[int]} }
        template_path: テンプレートExcelのパス（.xlsm）

    Returns:
        出力ファイルのパス文字列（例: output_answered_20250224.xlsx）
    """
    # テンプレートを読み込む（VBAマクロは保持しない。出力は.xlsx）
    wb = openpyxl.load_workbook(template_path, keep_vba=False, data_only=True)
    ws = wb[SHEET_NAME]

    # E列にヘッダーを追記
    ws.cell(row=HEADER_ROW, column=COL_EVIDENCE).value = "根拠ページ"

    # 各行のQIDを見て回答を書き込む
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        qid_cell = ws.cell(row=row_num, column=COL_QID)
        qid = qid_cell.value
        if qid is None:
            continue

        try:
            qid_int = int(qid)
        except (ValueError, TypeError):
            continue

        if qid_int not in answers:
            continue

        entry = answers[qid_int]
        answer_text    = entry.get("answer", "")
        evidence_pages = entry.get("evidence_pages", [])

        # D列: 回答テキスト
        ws.cell(row=row_num, column=COL_ANSWER).value = answer_text

        # E列: 根拠ページ（カンマ区切り、空の場合は空文字）
        if evidence_pages:
            pages_str = ", ".join(str(p) for p in evidence_pages)
            ws.cell(row=row_num, column=COL_EVIDENCE).value = f"p.{pages_str}"
        else:
            ws.cell(row=row_num, column=COL_EVIDENCE).value = ""

    # 出力ファイル名（日付付き）
    date_str    = datetime.now().strftime("%Y%m%d")
    output_path = f"output_answered_{date_str}.xlsx"

    wb.save(output_path)
    print(f"✅ Excel書き込み完了: {output_path}  ({len(answers)}問)")
    return output_path


# ===================================================
# 動作確認（単体テスト用）
# ===================================================
if __name__ == "__main__":
    # ダミーデータで書き込みテスト
    dummy_answers = {
        i: {
            "answer": f"Q{i} のサンプル回答です。",
            "evidence_pages": [i * 2, i * 2 + 1],
        }
        for i in range(1, 106)
    }
    path = write_answers_to_excel(dummy_answers)
    print(f"出力先: {path}")
